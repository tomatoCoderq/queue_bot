import os.path
from os import getenv
from loguru import logger


from app.utils.database import database
from openpyxl import Workbook, load_workbook


def create_idt_name_map() -> dict:
    database.execute("""
    SELECT requests_queue.idt, students.name, students.surname FROM requests_queue INNER JOIN students
    ON requests_queue.idt = students.idt
    """)

    idt_name_map = {row[0]: (row[1], row[2]) for row in database.cursor.fetchall()}

    return idt_name_map


class Excel:
    def __init__(self):
        self.file_name = getenv("FILE")
        if os.path.exists(self.file_name):
            self.workbook = load_workbook(self.file_name)
            self.sheet = self.workbook.active
            logger.success(f"Loaded and activated curren worksheet from {self.file_name}")
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append([x[1] for x in database.fetchall_multiple("PRAGMA table_info(requests_queue)")])
            logger.success(f"Created new file workbook and worksheet in {self.file_name}. Added row with titles")

    def write(self, execute):
        idt_name_map = create_idt_name_map()

        for row in execute:
            row_list = list(row)
            row_list[1] = f"{idt_name_map[row_list[1]][0]} {idt_name_map[row_list[1]][1]}"
            self.sheet.append(row_list)

        self.workbook.save(self.file_name)
        logger.success("Added new rows and saved all data")
