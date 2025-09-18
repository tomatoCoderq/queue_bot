import os.path
from os import getenv
from loguru import logger


from sqlalchemy import text
from openpyxl import Workbook, load_workbook

from src.storages.sql.dependencies import database
from src.storages.sql.models import StudentModel
from openpyxl import Workbook, load_workbook


def create_idt_name_map() -> dict:
    records = database.fetch_all(
        """
        SELECT DISTINCT requests_queue.idt AS idt, students.name, students.surname
        FROM requests_queue
        INNER JOIN students ON requests_queue.idt = students.idt
        """,
        model=StudentModel,
    )
    return {record.idt: (record.name, record.surname) for record in records}


class Excel:
    def __init__(self):
        self.file_name = getenv("FILE")
        if os.path.exists(self.file_name):
            self.workbook = load_workbook(self.file_name)
            self.sheet = self.workbook.active
            logger.success(f"Loaded and activated curren worksheet from {self.file_name}")
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(
                [x[1] for x in database.fetchall_multiple(text("PRAGMA table_info(requests_queue)"))]
            )
            logger.success(f"Created new file workbook and worksheet in {self.file_name}. Added row with titles")

    def write(self, execute):
        idt_name_map = create_idt_name_map()

        for row in execute:
            row_list = list(row)
            row_list[1] = f"{idt_name_map[row_list[1]][0]} {idt_name_map[row_list[1]][1]}"
            self.sheet.append(row_list)

        self.workbook.save(self.file_name)
        logger.success("Added new rows and saved all data")
